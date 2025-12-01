from __future__ import annotations

from dataclasses import dataclass
from typing import Dict, List, Tuple
from datetime import datetime, date, time

from django.db import transaction

from openpyxl import load_workbook

from accounts.models import Collaborator
from metrics.models import MetricType, MetricRecord
from .models import UploadBatch


@dataclass
class RowResult:
    ok: bool
    reason: str = ""


# ---------- utilidades seguras ----------

def _norm(s) -> str:
    """Normaliza qualquer valor para comparação de cabeçalho."""
    try:
        return str(s).strip().lower()
    except Exception:
        return ""


def _parse_date(value) -> date | None:
    """Aceita datetime/date do Excel ou string em YYYY-MM-DD / DD/MM/YYYY / DD-MM-YYYY."""
    if value is None or value == "":
        return None
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    if isinstance(value, str):
        s = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
    return None


def _excel_fraction_day_to_minutes(x: float) -> float:
    """Converte fração do dia do Excel para minutos (0.5 dia = 720 min)."""
    return float(x) * 24.0 * 60.0


def _hhmmss_to_minutes(s: str) -> float | None:
    """Converte 'HH:MM:SS' ou 'H:MM' para minutos."""
    try:
        parts = s.strip().split(":")
        if len(parts) == 2:
            h, m = int(parts[0]), int(parts[1])
            sec = 0
        elif len(parts) == 3:
            h, m, sec = int(parts[0]), int(parts[1]), int(parts[2])
        else:
            return None
        return h * 60 + m + sec / 60.0
    except Exception:
        return None


def _looks_like_time_metric(metric: MetricType) -> bool:
    """
    Heurística simples para detectar métrica 'de tempo' (para converter para minutos).
    Ajuste conforme seu cadastro de MetricType.
    """
    code = (metric.code or "").lower()
    unit = (metric.unit or "").lower()
    name = (metric.name or "").lower()

    hints = ("time", "tempo", "hh:mm", "hhmm", "ti", "duracao", "duração", "sla")
    unit_hints = ("min", "minuto", "minutos", "hora", "horas", "h")

    return any(h in code for h in hints) or any(h in name for h in hints) or any(u in unit for u in unit_hints)


def _parse_value(metric: MetricType, value) -> float | None:
    """
    Converte o 'valor' para float.
    - Para métricas de tempo, converte para MINUTOS:
      * datetime.time -> minutos
      * string 'HH:MM(:SS)' -> minutos
      * número (fração do dia do Excel) -> minutos
    - Para demais métricas:
      * número -> float
      * string com vírgula -> float
    """
    if value is None or value == "":
        return None

    if _looks_like_time_metric(metric):
        # datetime.time
        if isinstance(value, time):
            return value.hour * 60 + value.minute + value.second / 60.0
        # datetime/datetime64 com parte de hora?
        if isinstance(value, datetime):
            return value.hour * 60 + value.minute + value.second / 60.0
        # numeric (Excel armazena tempo como fração do dia)
        if isinstance(value, (int, float)):
            x = float(value)
            # se for claramente um horário (0 <= x < 1) trata como fração do dia
            if 0.0 <= x < 1.0:
                return _excel_fraction_day_to_minutes(x)
            # se já veio em minutos (número grande), mantém
            return x
        # string 'HH:MM' / 'HH:MM:SS'
        if isinstance(value, str):
            s = value.strip()
            mm = _hhmmss_to_minutes(s)
            if mm is not None:
                return mm
            # tenta número com vírgula/pontos
            s2 = s.replace(" ", "").replace(".", "").replace(",", ".")
            try:
                return float(s2)
            except ValueError:
                return None
        return None

    # Métrica comum (não-tempo)
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        s = value.strip().replace(" ", "")
        s = s.replace(".", "").replace(",", ".")
        try:
            return float(s)
        except ValueError:
            return None
    return None


# ---------- leitura da planilha ----------

_HEADER_ALIASES = {
    "colaborador_id": {"colaborador_id", "colaborador", "id_colaborador", "matricula", "matrícula", "codigo", "cod_colaborador", "cpf"},
    "data": {"data", "date", "dia"},
    "valor": {"valor", "value", "resultado", "indice", "índice", "pontuacao", "pontuação", "tr", "tempo"},
}

def _map_header_indices(header_cells) -> Tuple[Dict[str, int], Dict]:
    norm_names = [_norm(c) for c in header_cells]
    name_to_idx = {norm: i for i, norm in enumerate(norm_names)}

    idx_map: Dict[str, int] = {}
    for canonical, aliases in _HEADER_ALIASES.items():
        found_idx = None
        for alias in aliases:
            if alias in name_to_idx:
                found_idx = name_to_idx[alias]
                break
        if found_idx is not None:
            idx_map[canonical] = found_idx

    required = set(_HEADER_ALIASES.keys())
    if not required.issubset(idx_map.keys()):
        return {}, {
            "error": "Cabeçalho inválido",
            "found": norm_names,
            "expected": {k: sorted(list(v)) for k, v in _HEADER_ALIASES.items()},
        }
    return idx_map, {}


def _read_rows_from_workbook(uploaded_file, metric: MetricType) -> Tuple[List[Dict], Dict]:
    uploaded_file.seek(0)
    wb = load_workbook(uploaded_file, data_only=True, read_only=True)
    ws = wb.active  # primeira aba

    header_cells = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    idx_map, header_err = _map_header_indices(header_cells)
    if header_err:
        return [], header_err

    rows: List[Dict] = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            cid_raw = row[idx_map["colaborador_id"]] if len(row) > idx_map["colaborador_id"] else None
            d_raw = row[idx_map["data"]] if len(row) > idx_map["data"] else None
            v_raw = row[idx_map["valor"]] if len(row) > idx_map["valor"] else None

            d = _parse_date(d_raw)
            v = _parse_value(metric, v_raw)

            rows.append({
                "excel_row": r_idx,
                "colaborador_id": (str(cid_raw).strip() if cid_raw is not None else ""),
                "date": d,
                "value": v,
            })
        except Exception:
            rows.append({
                "excel_row": r_idx,
                "colaborador_id": "",
                "date": None,
                "value": None,
            })

    return rows, {}
# ---------- import principal ----------

def import_xlsx(metric: MetricType, uploaded_file, user) -> Tuple[bool, Dict]:
    """
    Importa uma planilha Excel (.xlsx ou .xls) criando/atualizando registros.
    Upsert por (colaborador, métrica, data). Salva FK do lote em source_batch_id.
    """
    rows, header_err = _read_rows_from_workbook(uploaded_file, metric)
    if header_err:
        return False, header_err

    created = 0
    updated = 0
    errors: List[Dict] = []

    batch = UploadBatch.objects.create(
        user=user,
        metric_type=metric,
        original_filename=getattr(uploaded_file, "name", "upload.xlsx"),
        report={}
    )

    with transaction.atomic():
        for r in rows:
            cid = (r.get("colaborador_id") or "").strip()
            d: date | None = r.get("date")
            v = r.get("value")

            if not cid or not d or v is None:
                errors.append({
                    "row": r.get("excel_row"),
                    "reason": "Linha incompleta (colaborador_id/data/valor)"
                })
                continue

            try:
                collab = Collaborator.objects.get(colaborador_id=cid)
            except Collaborator.DoesNotExist:
                errors.append({
                    "row": r.get("excel_row"),
                    "reason": f"colaborador_id '{cid}' não encontrado"
                })
                continue

            # Upsert + FK do batch via _id
            obj, was_created = MetricRecord.objects.update_or_create(
                collaborator=collab,
                metric_type=metric,
                date=d,
                defaults={
                    "value": v,
                    "source_batch_id": batch.id,  # <-- chave da correção de FK
                },
            )
            created += int(was_created)
            updated += int(not was_created)

    batch.report = {
        "imported": created + updated,
        "created": created,
        "updated": updated,
        "errors": errors,
    }
    batch.save(update_fields=["report"])

    ok = len(errors) == 0
    return ok, batch.report
